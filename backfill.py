"""
Backfill the archive from the workbook's daily tabs (e.g. 5.29 ... 6.23).

Each dated tab holds that day's CIF and barge freight. We read those plus the
delivery-window / contract calendar and write one snapshot per date into the
same store the app uses (SQLite by default, or Postgres if DATABASE_URL set).

Run:  python backfill.py "JSA FOB Sheet June 2026.xlsx"  [year]
"""
import re
import sys
import datetime as dt

import openpyxl

import db
import fob_model as M

# Row of each commodity's CIF entry, and each freight reach's entry row
# (the corn section — freight is shared across commodities).
CIF_ROW = {"Corn": 7, "Soybeans": 63, "Wheat": 118}
FREIGHT_ROW = {
    "Lower Miss": 9, "Davenport South": 15, "McGregor South": 19,
    "Upper Miss": 22, "Ohio": 25, "STL": 30, "IL": 33,
}
COLS = list(range(4, 12))   # D..K = the 8 month columns


def _num(v):
    return v if isinstance(v, (int, float)) else None


def backfill(path, year):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    db.init_db()
    saved = []
    for name in wb.sheetnames:
        if not re.fullmatch(r"\d{1,2}\.\d{1,2}", name.strip()):
            continue
        mo, day = (int(x) for x in name.strip().split("."))
        as_of = dt.date(year, mo, day).isoformat()
        ws = wb[name]

        cif = {c: {M.MONTHS[i]: _num(ws.cell(CIF_ROW[c], col).value)
                   for i, col in enumerate(COLS)}
               for c in M.COMMODITIES}
        freight = {r: {M.MONTHS[i]: _num(ws.cell(FREIGHT_ROW[r], col).value)
                       for i, col in enumerate(COLS)}
                   for r in FREIGHT_ROW}
        calendar = {c: list(zip(M.MONTHS, M.CONTRACTS[c])) for c in M.COMMODITIES}

        n_cif, n_frt = db.save_snapshot(as_of, cif, freight, calendar)
        saved.append((as_of, n_cif, n_frt))
        print(f"  {name:>6} -> {as_of}  ({n_cif} CIF, {n_frt} freight)")

    print(f"\nBackfilled {len(saved)} dates. Archive now: {db.list_dates()}")


if __name__ == "__main__":
    p = sys.argv[1] if len(sys.argv) > 1 else "JSA FOB Sheet June 2026.xlsx"
    y = int(sys.argv[2]) if len(sys.argv) > 2 else 2026
    backfill(p, y)
