r"""
Backfill the river archive from the CGB daily "Bid Sheet" workbooks
(…\San Antonio\CGB\MMDDYY.xlsx). Each file's "Bid Sheet" tab is a CIF NOLA +
barge-freight forward curve by month - the same shape the portal's paste parser
reads - so we pull CIF (Corn/Soybeans/Wheat) and barge freight per river segment
and store them; the portal recomputes FOB from those.

Weekly sampling: one file per week, the Wednesday's file if it exists, else the
nearest file within +/-`WINDOW` days. Existing archived dates are left untouched
(the JSA workbook data stays authoritative there) - only new dates are written.

Fixed Bid Sheet layout (validated across the range):
  col A       month label (TW/NW spot rows skipped)
  cols B..I   barge freight: ILL OHIO L-OHIO MM CITIES STL MTCT ARK
  M/N O/P Q/R S/T   CIF + contract letter for CORN MILO BEANS WHEAT (cents)

Usage:
  python import_cgb_history.py                     # DRY RUN (default range)
  python import_cgb_history.py --commit
  python import_cgb_history.py 2022-09-01 2023-08-31 --commit
"""
import os
import re
import sys
import glob
import shutil
import tempfile
import datetime as dt
import warnings

warnings.filterwarnings("ignore")
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"),
                override=True)
except ImportError:
    pass

import openpyxl
import db
import paste_parse as PP

CGB_ROOT = os.environ.get("CGB_ROOT") or (
    r"C:\Users\KoltenPostin\John Stewart and Associates\JSA - Documents"
    r"\San Antonio\CGB")
WINDOW = 6                                   # accept a file within N days of the Wed

# Bid Sheet columns -> canonical freight regions (skip L-Ohio col D and Ark col I).
FREIGHT_COLS = {
    2: ["IL"], 3: ["Ohio"], 5: ["Davenport South", "McGregor South"],
    6: ["Upper Miss"], 7: ["STL"], 8: ["Lower Miss"],
}
# CIF value col -> (commodity, contract-letter col). Skip Milo (col O).
CIF_COLS = {13: ("Corn", 14), 17: ("Soybeans", 18), 19: ("Wheat", 20)}
_PREFIX = {"Corn": "C", "Soybeans": "S", "Wheat": "W"}

# Older .xls format (2012-2015): data on 'Sheet1', freight month in col A and
# CIF month in col K, with the CIF block shifted one left of the .xlsx layout.
# Columns below are 0-indexed for xlrd.
FREIGHT_COLS_XLS = {1: ["IL"], 2: ["Ohio"], 4: ["Davenport South", "McGregor South"],
                    5: ["Upper Miss"], 6: ["STL"], 7: ["Lower Miss"]}
CIF_COLS_XLS = {11: ("Corn", 12), 15: ("Soybeans", 16), 17: ("Wheat", 18)}


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _month(label):
    """Month key from a label, folding split-half rows (FH/LH Feb -> Feb)."""
    return PP._MONTHS.get(re.sub(r"^(FH|LH)\s+", "", str(label or "").strip().upper()))


def index_files():
    """{date: path} for every MMDDYY.xlsx under the CGB root, including the
    marketing-year subfolders (06-07 … 17-18) that hold the older files. A root
    copy wins over a subfolder copy if the same date appears in both."""
    def _score(path):                        # prefer .xlsx over .xls, root over sub
        return (path.lower().endswith(".xlsx"), os.path.dirname(path) == CGB_ROOT)

    out = {}
    for pat in ("*.xlsx", "*.xls"):
        for f in glob.glob(os.path.join(CGB_ROOT, "**", pat), recursive=True):
            m = re.fullmatch(r"(\d{2})(\d{2})(\d{2})\.xlsx?", os.path.basename(f), re.I)
            if not m:
                continue
            try:
                d = dt.date(2000 + int(m.group(3)), int(m.group(1)), int(m.group(2)))
            except ValueError:
                continue
            if d not in out or _score(f) > _score(out[d]):
                out[d] = f
    return out


def wednesday_files(files, lo, hi):
    """Pick one file per week (Wed, else nearest within WINDOW). -> {date: path}."""
    in_range = [d for d in files if lo <= d <= hi]
    picks = {}
    w = lo
    while w <= hi:
        if w.weekday() == 2:                 # Wednesday
            cand = min(in_range, key=lambda d: (abs((d - w).days), d),
                       default=None)
            if cand and abs((cand - w).days) <= WINDOW:
                picks[cand] = files[cand]    # key by the file's real date
        w += dt.timedelta(days=1)
    return picks


def _finish(cif, freight, cal):
    cif = {c: mv for c, mv in cif.items() if mv}
    return (cif, freight, {c: v for c, v in cal.items() if v}) if cif else None


def _parse_xlsx(path):
    """Newer .xlsx 'Bid Sheet' layout (2016+). -> (cif, freight, calendar)."""
    tmp = os.path.join(tempfile.gettempdir(), "cgb_" + os.path.basename(path))
    shutil.copy2(path, tmp)
    try:
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
        if "Bid Sheet" not in wb.sheetnames:
            return None
        ws = wb["Bid Sheet"]
        cif = {c: {} for c in _PREFIX}
        cal = {c: [] for c in _PREFIX}
        freight = {}
        for r in range(3, 16):               # A3:T15 data rows
            mon = _month(ws.cell(r, 1).value)
            if not mon:                      # TW/NW spot rows and blanks
                continue
            for col, (cm, lc) in CIF_COLS.items():
                v = _num(ws.cell(r, col).value)
                if v is not None and v != 0 and mon not in cif[cm]:
                    cif[cm][mon] = round(v / 100.0, 4)
                    let = ws.cell(r, lc).value
                    if let:
                        cal[cm].append((mon, _PREFIX[cm] + str(let).strip()))
            for col, regs in FREIGHT_COLS.items():
                v = _num(ws.cell(r, col).value)
                if v is not None and v != 0:
                    for rg in regs:
                        freight.setdefault(rg, {}).setdefault(mon, round(v, 4))
        wb.close()
        return _finish(cif, freight, cal)
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass


def _parse_xls(path):
    """Older .xls 'Sheet1' layout (2012-2015): freight month col A, CIF month
    col K, CIF block one column left of the .xlsx layout."""
    import xlrd
    tmp = os.path.join(tempfile.gettempdir(), "cgbx_" + os.path.basename(path))
    shutil.copy2(path, tmp)
    try:
        wb = xlrd.open_workbook(tmp)
        if "Sheet1" not in wb.sheet_names():
            return None
        sh = wb.sheet_by_name("Sheet1")
        cif = {c: {} for c in _PREFIX}
        cal = {c: [] for c in _PREFIX}
        freight = {}
        for r in range(2, 15):               # 0-indexed rows 3..15
            fm = _month(sh.cell_value(r, 0)) if sh.ncols > 0 else None
            if fm:
                for col, regs in FREIGHT_COLS_XLS.items():
                    if col < sh.ncols:
                        v = _num(sh.cell_value(r, col))
                        if v is not None and v != 0:
                            for rg in regs:
                                freight.setdefault(rg, {}).setdefault(fm, round(v, 4))
            cm = _month(sh.cell_value(r, 10)) if sh.ncols > 10 else None
            if cm:
                for col, (cmd, lc) in CIF_COLS_XLS.items():
                    if col < sh.ncols:
                        v = _num(sh.cell_value(r, col))
                        if v is not None and v != 0 and cm not in cif[cmd]:
                            cif[cmd][cm] = round(v / 100.0, 4)
                            let = sh.cell_value(r, lc) if lc < sh.ncols else None
                            if let:
                                cal[cmd].append((cm, _PREFIX[cmd] + str(let).strip()))
        return _finish(cif, freight, cal)
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass


def parse_bidsheet(path):
    """Dispatch by format. A few .xls-extension files are actually xlsx, so try
    the extension-appropriate reader first and fall back to the other."""
    order = ((_parse_xls, _parse_xlsx) if path.lower().endswith(".xls")
             else (_parse_xlsx, _parse_xls))
    for fn in order:
        try:
            res = fn(path)
            if res:
                return res
        except Exception:
            continue
    return None


def run(lo, hi, commit):
    files = index_files()
    picks = wednesday_files(files, lo, hi)
    existing = {dt.date.fromisoformat(s) for s in db.list_dates()}
    todo = sorted(d for d in picks if d not in existing)
    skip = sorted(d for d in picks if d in existing)

    print(f"CGB Bid Sheets - {lo} -> {hi}")
    print(f"  files matched (weekly): {len(picks)}  "
          f"({min(picks)} -> {max(picks)})" if picks else "  no files matched")
    print(f"  already archived (kept): {len(skip)}")
    print(f"  NEW to write: {len(todo)}"
          + (f"  ({todo[0]} -> {todo[-1]})" if todo else ""))

    parsed = {}
    for d in todo:
        res = parse_bidsheet(picks[d])
        if res:
            parsed[d] = res
        else:
            print(f"    {d}: parse failed - skipped")
    for d in list(parsed)[:2] + list(parsed)[-1:]:
        cif, frt, cal = parsed[d]
        print(f"    {d}: Corn CIF={cif.get('Corn')}  regions={sorted(frt)}")

    if not commit:
        print(f"\nDRY RUN - nothing written. {len(parsed)} snapshots ready. "
              "Re-run with --commit.")
        return
    if not db._is_postgres():
        print("\nREFUSING: DATABASE_URL not set (would hit local SQLite). Aborting.")
        sys.exit(1)

    db.init_db()
    n = 0
    for d, (cif, frt, cal) in parsed.items():
        # Retry — a pooled connection can drop mid-run on a long write burst.
        for attempt in range(4):
            try:
                db.save_snapshot(d.isoformat(), cif, frt, cal)
                break
            except Exception:
                if attempt == 3:
                    raise
                import time
                time.sleep(3)
        n += 1
    now = db.list_dates()
    print(f"\nCommitted {n} snapshots. Archive now {len(now)} dates, "
          f"earliest {min(now)}.")


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    lo = dt.date.fromisoformat(args[0]) if len(args) > 0 else dt.date(2022, 9, 1)
    hi = dt.date.fromisoformat(args[1]) if len(args) > 1 else dt.date(2023, 8, 31)
    run(lo, hi, commit="--commit" in sys.argv)
