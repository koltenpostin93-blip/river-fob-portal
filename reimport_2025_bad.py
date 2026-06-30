"""
Re-import the 11 corrupt 2025 snapshots (late-Apr -> Jun) that the adaptive
parser botched because those workbooks' CBOT rows are #NAME? (Barchart add-in
not connected), leaving the column detection with nothing to latch onto.

These workbooks have a stable layout, so we read fixed rows/columns instead:
  month header  = CIF row - 3   (cols E..L = Spot + 7 months)
  CIF rows      = Corn 7, Soybeans 62, Wheat 117
  freight rows  = shared corn section (Lower Miss 9 ... IL 33)

Writes to whatever store db.py points at: local SQLite by default, or Postgres
when DATABASE_URL is set. save_snapshot replaces the existing (bad) rows for
each date, so this cleanly overwrites the junk.

    python reimport_2025_bad.py              # -> local SQLite
    $env:DATABASE_URL="..."; python reimport_2025_bad.py   # -> Postgres
"""
import os
import openpyxl
import db
import fob_model as M

SCRATCH = os.environ.get(
    "FOB2025_DIR",
    r"C:\Users\KOLTEN~1\AppData\Local\Temp\claude"
    r"\C--Users-KoltenPostin-OneDrive---John-Stewart-and-Associates-Desktop-Claude-Code"
    r"\7ac7d00b-ac05-43a0-933c-b3f1f62474d8\scratchpad")

# date -> (workbook stem, tab name)
JOBS = {
    "2025-04-23": ("April", "4-23-25"), "2025-04-28": ("April", "4-28-25"),
    "2025-04-30": ("April", "4-30-25"), "2025-05-07": ("May", "5-7-25"),
    "2025-05-14": ("May", "5-14-25"), "2025-05-21": ("May", "5-21-25"),
    "2025-05-28": ("May", "5-28-25"), "2025-06-04": ("June", "6-4-25"),
    "2025-06-11": ("June", "6-11-25"), "2025-06-18": ("June", "6-18-25"),
    "2025-06-25": ("June", "6-25-25"),
}
CIF_ROW = {"Corn": 7, "Soybeans": 62, "Wheat": 117}
CONTRACT_OFFSET = 2   # contract row sits 2 above the CIF row
HEADER_OFFSET = 3     # month header sits 3 above the CIF row
FREIGHT_ROW = {"Lower Miss": 9, "Davenport South": 15, "McGregor South": 19,
               "Upper Miss": 22, "Ohio": 25, "STL": 30, "IL": 33}
COLS = list(range(5, 13))   # E..L


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _labels(ws, row):
    out = []
    for c in COLS:
        v = ws.cell(row, c).value
        out.append(str(v).strip() if v is not None else "")
    return out


def reimport():
    db.init_db()
    cache = {}
    for as_of, (stem, tab) in JOBS.items():
        if stem not in cache:
            cache[stem] = openpyxl.load_workbook(
                os.path.join(SCRATCH, f"{stem}2025.xlsx"), data_only=True)
        ws = cache[stem][tab]

        cif, calendar = {}, {}
        for c in M.COMMODITIES:
            cr = CIF_ROW[c]
            months = _labels(ws, cr - HEADER_OFFSET)
            contracts = _labels(ws, cr - CONTRACT_OFFSET)
            row = {}
            cols = []
            for i, col in enumerate(COLS):
                m = months[i]
                if not m:
                    continue
                v = _num(ws.cell(cr, col).value)
                if v is not None:
                    row[m] = v
                cols.append((m, contracts[i] or None))
            cif[c] = row
            calendar[c] = cols

        # Freight is shared — read from the corn section, keyed by corn's months.
        corn_months = _labels(ws, CIF_ROW["Corn"] - HEADER_OFFSET)
        freight = {}
        for r, frow in FREIGHT_ROW.items():
            mv = {}
            for i, col in enumerate(COLS):
                m = corn_months[i]
                if not m:
                    continue
                v = _num(ws.cell(frow, col).value)
                if v is not None:
                    mv[m] = v
            freight[r] = mv

        n_cif, n_frt = db.save_snapshot(as_of, cif, freight, calendar)
        print(f"  {as_of} [{tab:8}] -> {n_cif} CIF, {n_frt} freight")

    print(f"\nRe-imported {len(JOBS)} dates into {db.backend_name()}.")


if __name__ == "__main__":
    reimport()
