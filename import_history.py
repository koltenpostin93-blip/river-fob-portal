"""
Import historical FOB sheets into the archive (CIF + barge freight + calendar).

Adaptive parser — handles the layout drift across years:
  * label column auto-detected (A or B)
  * data columns derived from the CBOT row's numeric cells (works for the
    2023-era Spot/Mar/Apr... set and the 2026-era June...Jan set alike), with a
    fallback to the month-header row when the CBOT row is #NAME? (Barchart
    add-in not connected — otherwise the column detection finds nothing)
  * CIF / freight / FOB rows + month & contract headers located by label text
  * freight-row name variants normalised ("IL Barge Freight" -> "IL")

Sampling (to keep the archive lean): every trading day within RECENT_DAYS of the
anchor date is kept; older dates are thinned to one per ISO week, preferring the
day closest to Wednesday.

Usage:
  python import_history.py            # DRY RUN — report only, writes nothing
  python import_history.py --commit   # actually write to the archive
"""
import os
import re
import sys
import glob
import warnings
import datetime as dt

warnings.filterwarnings("ignore")
import openpyxl

import db
import fob_model as M

FOB_ROOT = os.environ.get("FOB_ROOT") or (
    r"C:\Users\KoltenPostin\John Stewart and Associates"
    r"\JSA - Documents\St. Louis\JSA FOB Sheet")
ANCHOR = dt.date(2026, 6, 24)     # "today" for the recent-daily window
RECENT_DAYS = 14                  # keep daily within this many days of ANCHOR

MONTH_NAME = ("January February March April May June July August September "
              "October November December").split()
MONTHS_RE = re.compile(r"JSA FOB Sheet.*(" + "|".join(MONTH_NAME) + r")\s*(\d{4})", re.I)
DATED_TAB = re.compile(r"\d{1,2}[-.]\d{1,2}([-.]\d{2,4})?")

COMMODITY_ALIASES = {"corn": "Corn", "soybeans": "Soybeans", "soybean": "Soybeans",
                     "beans": "Soybeans", "wheat": "Wheat"}
# canonical freight regions we store (MTCT mirrors Lower Miss and is skipped)
REGION_ALIASES = {
    "lower miss": "Lower Miss", "davenport south": "Davenport South",
    "mcgregor south": "McGregor South", "upper miss": "Upper Miss",
    "ohio": "Ohio", "stl": "STL", "il": "IL",
}


def list_workbooks():
    out = []
    for f in glob.glob(os.path.join(FOB_ROOT, "**", "*.xlsx"), recursive=True):
        b = os.path.basename(f)
        if not MONTHS_RE.search(b):
            continue
        if "History" in f or "Client Files" in f:
            continue
        if "June 2026" in b:
            continue  # open in Excel + its daily tabs are already archived
        # Note: odd names (May 2026 V2, June 20251, February 20261, ...-PC) are
        # kept — true duplicates just upsert to the same dates, which is safe.
        out.append(f)
    return sorted(out)


def _tab_date(tab, wb_year, wb_month):
    """Parse a tab name into a date. Handles 'M-D-YY', 'M.D', 'M-D'."""
    parts = re.split(r"[-.]", tab.strip())
    try:
        mo, day = int(parts[0]), int(parts[1])
    except (ValueError, IndexError):
        return None
    if len(parts) >= 3 and parts[2]:
        yy = int(parts[2])
        year = 2000 + yy if yy < 100 else yy
    else:
        # infer year from the workbook, handling Dec/Jan wrap
        year = wb_year
        if wb_month == 12 and mo == 1:
            year += 1
        elif wb_month == 1 and mo == 12:
            year -= 1
    try:
        return dt.date(year, mo, day)
    except ValueError:
        return None


def _norm_region(label):
    s = label[:-len("Freight")].strip()
    if s.lower().endswith("barge"):
        s = s[:-len("barge")].strip()
    return REGION_ALIASES.get(s.lower())


_MNUM = {"jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
         "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
         "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9, "oct": 10,
         "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12}
_ABBR = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun", 7: "Jul",
         8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}


def _canon_month(label):
    """Normalize a column header to 'Spot' or 'Jan'..'Dec'; None for junk
    (numbers, blanks, TW/NW, half-month FH/LH)."""
    s = str(label).strip().lower().rstrip(".")
    if s == "spot":
        return "Spot"
    if s[:2] in ("fh", "lh") or s in ("tw", "nw", ""):
        return None
    return _ABBR.get(_MNUM.get(s))


def parse_tab(ws):
    """Return (cif, freight, calendar, futures) for one dated worksheet, or None.

    futures = {commodity: {month: cbot_price}} — the CBOT row values, present
    only when the workbook was saved with the Barchart add-in resolved (else the
    cells read #NAME?/#NUM! and futures come back empty)."""
    # label column: A or B by where "FOB Barge" appears most
    a = sum(1 for r in range(1, 60) if "FOB Barge" in str(ws.cell(r, 1).value or ""))
    b = sum(1 for r in range(1, 60) if "FOB Barge" in str(ws.cell(r, 2).value or ""))
    lc = 1 if a >= b else 2
    if max(a, b) == 0:
        return None

    rows = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, lc).value
        if isinstance(v, str) and v.strip():
            rows.setdefault(r, v.strip())

    # commodity sections: header row -> commodity
    sections = []
    for r, lbl in rows.items():
        key = COMMODITY_ALIASES.get(lbl.strip().lower())
        if key:
            sections.append((r, key))
    sections.sort()
    if not sections:
        return None
    bounds = [(sections[i][0], sections[i][1],
               sections[i + 1][0] if i + 1 < len(sections) else ws.max_row + 1)
              for i in range(len(sections))]

    cif, freight, calendar, futures = {}, {}, {}, {}
    for start, commodity, end in bounds:
        cbot_r = next((r for r in range(start, end)
                       if str(ws.cell(r, lc).value or "").strip().upper() == "CBOT"), None)
        if not cbot_r:
            continue
        # The month header sits 2 rows above CBOT, the contract row 1 above.
        header_r, contract_r = cbot_r - 2, cbot_r - 1

        def _collect(candidate_cols, contiguous=False):
            """Keep columns whose header normalises to a real month; drop
            duplicates (leftmost = the true forward-curve column). With
            contiguous=True, stop at the first gap once the curve has begun so a
            far-right month-labelled panel can't be swept in."""
            dcols, mons, cons, seen = [], [], [], set()
            started = False
            for c in candidate_cols:
                lbl = _canon_month(ws.cell(header_r, c).value)
                if lbl is None or lbl in seen:
                    if contiguous and started:
                        break
                    continue
                started = True
                seen.add(lbl)
                dcols.append(c)
                mons.append(lbl)
                cons.append(str(ws.cell(contract_r, c).value or "").strip())
            return dcols, mons, cons

        # Primary: columns where the CBOT row is numeric (Barchart connected) —
        # this also screens out stray numeric panels to the right.
        cbot_cols = [c for c in range(lc + 1, ws.max_column + 1)
                     if isinstance(ws.cell(cbot_r, c).value, (int, float))]
        data_cols, months, contracts = _collect(cbot_cols)
        # Header-row detection (contiguous run) — the reliable source when the
        # CBOT row is #NAME?/#NUM!/blank (add-in not connected) or only partly
        # resolved. Use it whenever it reveals a longer curve than CBOT did; on a
        # fully-connected sheet the two agree, so good workbooks are unchanged.
        h_cols, h_months, h_contracts = _collect(
            range(lc + 1, ws.max_column + 1), contiguous=True)
        if len(h_cols) > len(data_cols):
            data_cols, months, contracts = h_cols, h_months, h_contracts
        if not data_cols:
            continue
        calendar[commodity] = list(zip(months, contracts))

        # CBOT futures row (only real when Barchart was resolved on save).
        # Normalise to $/bu: some commodities' cells are in cents (e.g. soybeans
        # 1108.75) while others are already dollars (corn 4.07) — a value over
        # 100 is cents, so /100.
        fut = {months[i]: _fut_dollars(ws.cell(cbot_r, c).value)
               for i, c in enumerate(data_cols)}
        # Only keep the CBOT row when every value is a plausible grain price
        # ($/bu). A disconnected Barchart add-in caches junk (e.g. 25/26 for
        # every month) — skip it so the date archives CIF+freight only rather
        # than polluting the sheet's CBOT/spreads with garbage.
        vals = [v for v in fut.values() if v is not None]
        if vals and all(1.5 <= v <= 20 for v in vals):
            futures[commodity] = fut

        for r in range(start, end):
            lbl = str(ws.cell(r, lc).value or "").strip()
            if lbl.upper() == "CIF":
                cif[commodity] = {months[i]: _num(ws.cell(r, c).value)
                                  for i, c in enumerate(data_cols)}
            elif lbl.endswith("Freight"):
                reg = _norm_region(lbl)
                if reg and reg not in freight:   # first occurrence wins
                    freight[reg] = {months[i]: _num(ws.cell(r, c).value)
                                    for i, c in enumerate(data_cols)}
    if not cif:
        return None
    return cif, freight, calendar, futures


def spreads_from(futures, calendar):
    """Build {commodity: [(label, value), ...]} from captured CBOT futures and
    the tab's own contract chain — so archived days carry spreads too."""
    out = {}
    for c, fut in (futures or {}).items():
        cols = (calendar or {}).get(c) or []
        contracts = [ct for _m, ct in cols]
        months = [m for m, _ct in cols]
        vals = M.spreads_from_futures(c, fut, contracts=contracts, months=months)
        labels = M.spread_labels_for(c, contracts)
        pairs = [(l, v) for l, v in zip(labels, vals) if v is not None]
        if pairs:
            out[c] = pairs
    return out


def _fut_dollars(v):
    """CBOT price as $/bu — cells over 100 are quoted in cents, so /100."""
    v = _num(v)
    if v is None:
        return None
    return round(v / 100.0, 4) if abs(v) > 100 else v


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _weekly_sample(dates):
    """Keep daily within RECENT_DAYS of ANCHOR; else one/week nearest Wednesday."""
    keep = set(d for d in dates if (ANCHOR - d).days <= RECENT_DAYS)
    weeks = {}
    for d in dates:
        if d in keep:
            continue
        weeks.setdefault(d.isocalendar()[:2], []).append(d)
    for group in weeks.values():
        keep.add(min(group, key=lambda d: (abs(d.weekday() - 2), d)))
    return sorted(keep)


def run(commit):
    parsed = {}      # date -> (cif, freight, calendar)
    failures = []
    wbs = list_workbooks()
    print(f"Scanning {len(wbs)} workbooks...\n")
    for f in wbs:
        b = os.path.basename(f)
        m = MONTHS_RE.search(b)
        wb_month = MONTH_NAME.index(m.group(1).capitalize()) + 1
        wb_year = int(m.group(2))
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            failures.append((b, f"open: {e}"))
            continue
        n = 0
        for tab in wb.sheetnames:
            if not DATED_TAB.fullmatch(tab.strip()):
                continue
            d = _tab_date(tab, wb_year, wb_month)
            if not d:
                continue
            res = parse_tab(wb[tab])
            if res:
                parsed[d] = res
                n += 1
        print(f"  {b:42s} {n:3d} tabs")
    if not parsed:
        print("\nNo data parsed."); return

    selected = _weekly_sample(sorted(parsed))
    daily = [d for d in selected if (ANCHOR - d).days <= RECENT_DAYS]
    print(f"\nParsed {len(parsed)} trading days "
          f"({min(parsed)} -> {max(parsed)}).")
    print(f"After sampling: {len(selected)} snapshots "
          f"({len(daily)} daily-recent, {len(selected) - len(daily)} weekly).")
    locs = sorted({c for d, (cif, fr, cal, fut) in parsed.items() for c in cif})
    print(f"Commodities seen: {locs}")
    if failures:
        print("\nFailures:")
        for b, e in failures:
            print(f"  {b}: {e}")

    if not commit:
        print("\nDRY RUN — nothing written. Re-run with --commit to archive.")
        sample = selected[::max(1, len(selected) // 12)][:12]
        print("Sample of dates that would be saved:",
              [d.isoformat() for d in sample])
        return

    db.init_db()
    written = 0
    for d in selected:
        cif, freight, calendar, futures = parsed[d]
        # Retry each write — a pooled Postgres connection can be dropped mid-run
        # under a long burst of writes ("server closed the connection").
        for attempt in range(4):
            try:
                db.save_snapshot(
                    d.isoformat(), cif, freight, calendar,
                    futures=futures, spreads=spreads_from(futures, calendar))
                break
            except Exception as e:
                if attempt == 3:
                    raise
                import time
                time.sleep(3)
        written += 1
    print(f"\nCommitted {written} snapshots. Archive now spans "
          f"{min(db.list_dates())} -> {max(db.list_dates())} "
          f"({len(db.list_dates())} dates).")


if __name__ == "__main__":
    run(commit="--commit" in sys.argv)
