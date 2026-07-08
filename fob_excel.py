"""
Export the working FOB sheets to a single-worksheet .xlsx: Corn, then Soybeans,
then Wheat, stacked with one blank row between each. The worksheet is named the
as-of date (e.g. "7-7-26"). Values are written as real numbers with formats that
mirror the on-screen sheet (red negatives in parens, freight/% as percentages).

The app passes `sheets`: [{"commodity", "months", "rows"}], where each row is
(kind, label, cells); cells is a list aligned to the months (None for a
full-width section/banner row). Cell values are raw (float / str / None).
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK = "FF32373C"
BLUE = "FF0693E3"
GREY = "FFEEF1F4"
RED = "FFC00000"
GRIDC = "FFDDDDDD"

_thin = Side(style="thin", color=GRIDC)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# Number format per row kind (red negatives shown in parentheses, like the sheet).
NUMFMT = {
    "cbot": "0.0000",
    "cif": "0.00;[Red](0.00)",
    "cash": "0.00;[Red](0.00)",
    "fob": "0.00;[Red](0.00)",
    "topcarry": "0.00;[Red](0.00)",
    "spread": "0.0000;[Red](0.0000)",
    "freight": "0%",
    "carry": "0%",
}


def build_xlsx(as_of, sheets):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{as_of.month}-{as_of.day}-{as_of.year % 100}"

    n_months = len(sheets[0]["months"]) if sheets else 8
    ncol = n_months + 1
    r = 1
    for sheet in sheets:
        for kind, label, cells in sheet["rows"]:
            last_col = get_column_letter(ncol)

            if kind == "banner":
                ws.merge_cells(f"A{r}:{last_col}{r}")
                c = ws.cell(r, 1, label.upper())
                c.fill = PatternFill("solid", fgColor=BLUE)
                c.font = Font(bold=True, color="FFFFFFFF", size=12)
                c.alignment = Alignment(horizontal="center", vertical="center")
                r += 1
                continue

            if kind == "section" or cells is None:
                is_cash = label.lower().startswith("cash")
                ws.merge_cells(f"A{r}:{last_col}{r}")
                c = ws.cell(r, 1, label)
                c.fill = PatternFill("solid", fgColor=BLUE if is_cash else GREY)
                c.font = Font(bold=True,
                              color="FFFFFFFF" if is_cash else DARK, size=9)
                c.alignment = Alignment(
                    horizontal="center" if is_cash else "left", vertical="center")
                r += 1
                continue

            # label cell
            lab = ws.cell(r, 1, label)
            lab.border = BORDER
            if kind in ("cbot", "cif", "cash", "topcarry"):
                lab.font = Font(bold=True, color=DARK, size=9)
            elif kind == "freight":
                lab.font = Font(italic=True, color="FF555555", size=9)
            elif kind in ("months", "contracts"):
                lab.font = Font(bold=True, color="FFFFFFFF", size=9)
                lab.fill = PatternFill("solid", fgColor=DARK)
            else:
                lab.font = Font(color=DARK, size=9)

            header = kind in ("months", "contracts")
            for j, val in enumerate(cells):
                cell = ws.cell(r, 2 + j)
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if header:
                    cell.value = val
                    cell.font = Font(bold=True, color="FFFFFFFF", size=9)
                    cell.fill = PatternFill("solid", fgColor=DARK)
                elif val is None:
                    continue
                elif isinstance(val, str):          # e.g. a spread label
                    cell.value = val
                    cell.font = Font(italic=True, color="FF6B7280", size=9)
                else:
                    cell.value = val
                    cell.number_format = NUMFMT.get(kind, "0.00")
                    cell.font = Font(size=9)
            r += 1
        r += 1  # one blank row between commodities

    ws.column_dimensions["A"].width = 26
    for j in range(n_months):
        ws.column_dimensions[get_column_letter(2 + j)].width = 9.5
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B1"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
